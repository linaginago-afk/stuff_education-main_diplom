import csv
import io
import re
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import func
from sqlalchemy.orm import Session
from .. import schemas
from ..database import get_db
from ..deps import require_admin
from ..models import User
from ..security import get_password_hash

router = APIRouter(prefix="/api/admin/users", tags=["admin-users"])

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

HEADER_ALIASES = {
    "full_name": {"фио", "ф.и.о.", "full_name", "fullname", "name", "fio", "сотрудник"},
    "email": {"email", "e-mail", "mail", "почта", "электронная почта", "логин"},
    "password": {"пароль", "password", "pass"},
    "department": {"подразделение", "department", "отдел", "департамент"},
}


def _normalize_email(value: str) -> str:
    return value.strip().lower()


def _validate_email(value: str) -> str:
    email = _normalize_email(value)
    if not EMAIL_RE.match(email):
        raise ValueError("Некорректный email")
    return email


def _find_user_by_username(db: Session, username: str) -> User | None:
    return db.query(User).filter(func.lower(User.username) == username.lower()).first()


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _header_key(header: object) -> str | None:
    normalized = _normalize_header(header)
    for key, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return key
    return None


def _map_row(raw_row: dict[str, object], row_number: int) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for header, value in raw_row.items():
        key = _header_key(header)
        if key:
            mapped[key] = str(value or "").strip()

    required = ("full_name", "email", "password")
    missing = [field for field in required if not mapped.get(field)]
    if missing:
        raise ValueError(f"Не заполнены обязательные поля: {', '.join(missing)}")
    if len(mapped["password"]) < 4:
        raise ValueError("Пароль должен содержать минимум 4 символа")
    mapped["email"] = _validate_email(mapped["email"])
    mapped["department"] = mapped.get("department", "")
    return mapped


def _read_csv_rows(content: bytes) -> list[tuple[int, dict[str, object]]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать CSV: {last_error}")

    sample = text[:2048]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV-файл не содержит строку заголовков")
    return [
        (idx, row)
        for idx, row in enumerate(reader, start=2)
        if any(value not in (None, "") for value in row.values())
    ]


def _read_xlsx_rows(content: bytes) -> list[tuple[int, dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="Для импорта Excel установите пакет openpyxl") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise HTTPException(status_code=400, detail="Excel-файл не содержит строку заголовков")
    return [
        (idx, dict(zip(headers, row)))
        for idx, row in enumerate(rows, start=2)
        if any(value not in (None, "") for value in row)
    ]


def _read_import_rows(file: UploadFile, content: bytes) -> list[tuple[int, dict[str, object]]]:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(content)
    if suffix == ".xlsx":
        return _read_xlsx_rows(content)
    raise HTTPException(status_code=400, detail="Поддерживаются только файлы .csv и .xlsx")


@router.get("", response_model=list[schemas.UserOut])
def list_users(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    return db.query(User).order_by(User.id).all()


@router.post("", response_model=schemas.UserOut)
def create_user(payload: schemas.UserCreate, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    username = _validate_email(payload.username) if "@" in payload.username else payload.username.strip()
    if not username:
        raise HTTPException(status_code=400, detail="Email или логин не может быть пустым")
    if _find_user_by_username(db, username):
        raise HTTPException(status_code=400, detail="Пользователь с таким email или логином уже существует")
    user = User(
        full_name=payload.full_name,
        username=username,
        password_hash=get_password_hash(payload.password),
        role=payload.role,
        department=payload.department,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.post("/import", response_model=schemas.UserImportResult)
async def import_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пуст")

    rows = _read_import_rows(file, content)
    result = schemas.UserImportResult(created=0, skipped=0)
    seen_emails: set[str] = set()

    for row_number, raw_row in rows:
        email = None
        try:
            row = _map_row(raw_row, row_number)
            email = row["email"]
            if email in seen_emails:
                raise ValueError("Дубликат email внутри файла")
            seen_emails.add(email)
            if _find_user_by_username(db, email):
                raise ValueError("Пользователь с таким email уже существует")

            db.add(
                User(
                    full_name=row["full_name"],
                    username=email,
                    password_hash=get_password_hash(row["password"]),
                    role="employee",
                    department=row.get("department") or None,
                )
            )
            result.created += 1
        except ValueError as exc:
            result.skipped += 1
            result.errors.append(schemas.UserImportError(row=row_number, email=email, error=str(exc)))

    db.commit()
    return result


@router.put("/{user_id}", response_model=schemas.UserOut)
def update_user(
    user_id: int,
    payload: schemas.UserUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.role is not None:
        user.role = payload.role
    if payload.department is not None:
        user.department = payload.department
    if payload.password:
        user.password_hash = get_password_hash(payload.password)
    db.commit()
    db.refresh(user)
    return user


@router.delete("/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), _: User = Depends(require_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(user)
    db.commit()
    return {"detail": "Deleted"}
